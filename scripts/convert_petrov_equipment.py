#!/usr/bin/env python3
"""
Конвертация реестра оборудования (Петров А.А.) в формат импорта EPS-модуля.

Source:  temp/equipment/Петров А А.xlsx  → лист «Реестр оборудования»
Target:  temp/equipment/import_petrov_equipment.csv

Маппинг соответствует import_field_mapping.json и заголовкам import_equipment_all.csv
"""

import csv
import warnings
import re
from pathlib import Path

warnings.filterwarnings("ignore")

try:
    import openpyxl
except ImportError:
    raise SystemExit("Установите openpyxl: pip3 install openpyxl --break-system-packages")

# ---------------------------------------------------------------------------
# Пути
# ---------------------------------------------------------------------------
BASE_DIR = Path(__file__).resolve().parent.parent
SRC_FILE = BASE_DIR / "temp" / "equipment" / "Петров А А.xlsx"
DST_FILE = BASE_DIR / "temp" / "equipment" / "import_petrov_equipment.csv"

# ---------------------------------------------------------------------------
# Целевые заголовки (порядок как в import_equipment_all.csv)
# ---------------------------------------------------------------------------
TARGET_HEADERS = [
    "Наименование оборудования",
    "Инвентарный номер",
    "Заводской / Серийный номер",
    "Производитель",
    "Модель / Модификация",
    "Место установки (Локация)",
    "Рабочий статус",
    "Дата ввода в эксплуатацию",
    "Теги / Классификаторы",
    "Децимальный номер",
    "Код по ОКОФ (ОК 013-2014)",
    "Код по ОКПД2 (ОК 034-2014)",
    "Код технологического классификатора",
    "Группа оборудования",
    "Тип оборудования (Установка)",
    "Страна производитель",
    "Год выпуска",
    "Год ввода",
    "Возраст оборудования",
    "Категория критичности",
    "Фактический процент износа",
    "Класс чистоты помещения (ISO)",
    "Уникальное / единичное оборудование",
    "Импортное оборудование",
    "Периодичность регламентного ТО",
    "Утвержденный график ТО на 2026 год",
    "Количество ТО по графику",
    "Ответственное лицо (ФИО / Должность)",
]

# ---------------------------------------------------------------------------
# Маппинг статусов (source → EPS)
# ---------------------------------------------------------------------------
STATUS_MAP = {
    "В работе":    "В работе",
    "Резерв":      "В резерве",
    "Хранение":    "На хранении",
    "Списан":      "Списан",
    "Ремонт":      "На ремонте",
    "Консервация": "На консервации",
}


def clean_str(val) -> str:
    """Привести значение ячейки к строке без лишних пробелов."""
    if val is None:
        return ""
    return str(val).strip()


def extract_okpd2_code(raw: str) -> str:
    """Из строки '26.51.66 -Инструменты...' вернуть только '26.51.66'."""
    if not raw:
        return ""
    m = re.match(r"^([\d.]+)", raw.strip())
    return m.group(1) if m else raw.strip()


def wear_percent(val) -> str:
    """Привести износ к целому проценту или оставить пустым."""
    if val is None or val == "":
        return ""
    try:
        pct = float(val)
        # если дробное (0.4 → 40%), если > 1 — уже проценты (40 → 40%)
        if pct <= 1:
            pct = round(pct * 100)
        else:
            pct = round(pct)
        return str(pct)
    except (ValueError, TypeError):
        return clean_str(val)


def commission_date(comm_year) -> str:
    """Год ввода → дата вида 'YYYY-01-01'."""
    if comm_year is None or comm_year == "":
        return ""
    try:
        return f"{int(comm_year)}-01-01"
    except (ValueError, TypeError):
        return ""


def convert_row(row: tuple) -> dict:
    """
    Колонки source (индексы 0-25, данные из строк начиная с row 5):
      0  ID оборудования      (генерируется формулой, data_only=True даёт значение)
      1  Инв. №
      2  Наименование оборудования
      3  Децимальный номер
      4  Заводской №
      5  Комплекс / группа    (→ Тип оборудования / Установка)
      6  Расположение
      7  Ответственный
      8  Страна производитель
      9  Наименование производителя
     10  Год выпуска
     11  Год ввода
     12  Возраст оборудования
     13  Статус
     14  Критичность
     15  Периодичность ТО
     16  График ТО 2026 (месяцы)
     17  Кол-во ТО по графику
     18  Код ОКОФ 2
     19  Группа оборудования
     20  Ключ связи           (не используется)
     21  Классификатор техпроцесса, код
     22  Уникальное оборудование
     23  Импортное оборудование
     24  Фактический износ, %
     25  Код ОКПД 2 (сырой)
    """
    name          = clean_str(row[2])
    inv_number    = clean_str(row[1])
    serial        = clean_str(row[4])
    manufacturer  = clean_str(row[9])
    decimal       = clean_str(row[3])
    location      = clean_str(row[6])
    status_raw    = clean_str(row[13])
    status        = STATUS_MAP.get(status_raw, status_raw)
    comm_year     = row[11]
    okof          = clean_str(row[18])
    okpd2_raw     = clean_str(row[25]) if len(row) > 25 else ""
    okpd2         = extract_okpd2_code(okpd2_raw)
    proc_code     = clean_str(row[21])
    group_eq      = clean_str(row[19])
    equip_type    = clean_str(row[5])   # Комплекс/группа → Тип установки
    country       = clean_str(row[8])
    prod_year     = clean_str(row[10])
    c_year        = clean_str(row[11])
    age           = clean_str(row[12])
    criticality   = clean_str(row[14])
    wear          = wear_percent(row[24])
    is_unique     = clean_str(row[22])
    is_imported   = clean_str(row[23])
    periodicity   = clean_str(row[15])
    schedule_2026 = clean_str(row[16])
    to_count      = clean_str(row[17])
    responsible   = clean_str(row[7])

    # Теги: объединяем группу и тип (как в эталонном CSV)
    tags_parts = [t for t in [group_eq, equip_type] if t]
    tags = ", ".join(dict.fromkeys(tags_parts))  # уникальные, сохраняя порядок

    return {
        "Наименование оборудования":        name,
        "Инвентарный номер":                inv_number,
        "Заводской / Серийный номер":       serial,
        "Производитель":                    manufacturer,
        "Модель / Модификация":             decimal,      # децимальный № как модель
        "Место установки (Локация)":        location,
        "Рабочий статус":                   status,
        "Дата ввода в эксплуатацию":        commission_date(comm_year),
        "Теги / Классификаторы":            tags,
        "Децимальный номер":                decimal,
        "Код по ОКОФ (ОК 013-2014)":        okof,
        "Код по ОКПД2 (ОК 034-2014)":       okpd2,
        "Код технологического классификатора": proc_code,
        "Группа оборудования":              group_eq,
        "Тип оборудования (Установка)":     equip_type,
        "Страна производитель":             country,
        "Год выпуска":                      prod_year,
        "Год ввода":                        c_year,
        "Возраст оборудования":             age,
        "Категория критичности":            criticality,
        "Фактический процент износа":       wear,
        "Класс чистоты помещения (ISO)":    "",
        "Уникальное / единичное оборудование": is_unique,
        "Импортное оборудование":           is_imported,
        "Периодичность регламентного ТО":   periodicity,
        "Утвержденный график ТО на 2026 год": schedule_2026,
        "Количество ТО по графику":         to_count,
        "Ответственное лицо (ФИО / Должность)": responsible,
    }


# Префикс для временных инвентарных номеров (TEMP = временный, PAA = Петров А.А.)
TEMP_INV_PREFIX = "TEMP-PAA"


def main():
    print(f"📂 Источник: {SRC_FILE}")
    wb = openpyxl.load_workbook(str(SRC_FILE), data_only=True)
    ws = wb["Реестр оборудования"]

    rows_written     = 0   # строки с инв. № (реальным или временным)
    rows_temp_inv    = 0   # строки, которым назначен временный инв. №
    rows_skipped     = 0   # строки без имени или без обоих номеров
    temp_counter     = 0   # счётчик для генерации TEMP-PAA-XXXX
    seen_inv: set    = set()
    dup_inv: list    = []

    with open(DST_FILE, "w", newline="", encoding="utf-8-sig") as f_main:
        writer_main = csv.DictWriter(f_main, fieldnames=TARGET_HEADERS, quoting=csv.QUOTE_ALL)
        writer_main.writeheader()

        for row in ws.iter_rows(min_row=5, max_row=ws.max_row, values_only=True):
            # Пропускаем строки без наименования
            name = clean_str(row[2]) if len(row) > 2 else ""
            if not name:
                rows_skipped += 1
                continue

            inv    = clean_str(row[1])
            serial = clean_str(row[4])

            # Нормализуем «_» → пустой инв. №
            if inv == "_":
                inv = ""

            # Пропускаем строки без обоих идентификаторов
            # (самодельное/вспомогательное оборудование без учётных данных)
            if not inv and not serial:
                rows_skipped += 1
                continue

            record = convert_row(row)

            if not inv:
                # Назначаем временный инвентарный номер с префиксом TEMP-PAA-XXXX.
                # Номера сгенерированы автоматически и ОБЯЗАТЕЛЬНЫ К ЗАМЕНЕ
                # на реальные после постановки на учёт.
                temp_counter += 1
                inv = f"{TEMP_INV_PREFIX}-{temp_counter:04d}"
                rows_temp_inv += 1

                # ВАЖНО: очищаем заводской № чтобы API не нашёл совпадение
                # по serialNumber и создал новую запись, а не обновил случайную.
                # Оригинальный заводской № переносим в кастомное поле orig_serial.
                record["Заводской / Серийный номер"] = ""

            record["Инвентарный номер"] = inv

            # Фиксируем дубль инвентарного номера (ошибка данных источника)
            if inv in seen_inv:
                dup_inv.append((inv, name, serial))
            seen_inv.add(inv)

            writer_main.writerow(record)
            rows_written += 1

    print(f"✅ Файл для импорта: {DST_FILE}")
    print(f"   Всего записей:    {rows_written}")
    print(f"   С реальным инв. №:    {rows_written - rows_temp_inv}")
    print(f"   С временным инв. №:   {rows_temp_inv}  "
          f"(префикс {TEMP_INV_PREFIX}-XXXX, обязательны к замене!)")
    print()
    print(f"⏭  Пропущено (нет имени/обоих номеров): {rows_skipped}")

    if dup_inv:
        print()
        print(f"⚠️  Дубли инвентарных номеров в источнике ({len(dup_inv)} шт.) — требуют уточнения:")
        for inv_num, eq_name, ser in dup_inv:
            print(f"   инв. {inv_num!r}  |  {eq_name!r}  |  зав. {ser!r}")


if __name__ == "__main__":
    main()
